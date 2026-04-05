from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import zipfile
from datetime import UTC, date, datetime, time
from uuid import uuid4

from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from app.core.config import get_settings
from app.models.governance import Attraction, Dataset, Project, ProjectDataset, ProjectMember
from app.models.organization import Organization
from app.models.planner import Itinerary, ItineraryDay, ItineraryStop, ItineraryVersion
from app.models.rbac import UserRole
from app.models.user import User
from app.services.authorization import user_has_any_permission

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

EARTH_RADIUS_MILES = 3958.8
URBAN_DISTANCE_PORTION_MILES = 10.0
IMPORT_EXPORT_COLUMNS = [
    "day_number",
    "day_title",
    "day_notes",
    "day_urban_speed_mph_override",
    "day_highway_speed_mph_override",
    "stop_order",
    "attraction_id",
    "attraction_name",
    "attraction_city",
    "attraction_state",
    "start_time",
    "duration_minutes",
    "stop_notes",
]
REQUIRED_IMPORT_COLUMNS = ["day_number", "stop_order", "attraction_id", "start_time", "duration_minutes"]
SUPPORTED_IMPORT_FORMATS = {"csv", "xlsx"}
SYNC_PACKAGE_TYPE = "trailforge_planner_sync_package"
SYNC_PACKAGE_FORMAT_VERSION = "1.0"
SYNC_MANIFEST_PATH = "manifest.json"
SYNC_DATA_PATH = "data/itineraries.json"
SYNC_ASSETS_PATH = "assets/attractions.json"


class PlannerConflictError(Exception):
    pass


class PlannerValidationError(Exception):
    pass


class PlannerAuthorizationError(Exception):
    pass


class PlannerPayloadTooLargeError(Exception):
    pass


def _is_org_admin(db: Session, user: User) -> bool:
    return user_has_any_permission(db, user_id=user.id, required_permissions=("org.manage",))


def _is_planner(db: Session, user: User) -> bool:
    return user_has_any_permission(db, user_id=user.id, required_permissions=("itinerary.write",))


def _project_membership(db: Session, *, project_id: str, user_id: str) -> ProjectMember | None:
    return (
        db.execute(select(ProjectMember).where(ProjectMember.project_id == project_id, ProjectMember.user_id == user_id))
        .scalars()
        .first()
    )


def _project_for_user(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    user: User,
    require_edit: bool,
) -> Project | None:
    project = db.execute(select(Project).where(Project.id == project_id, Project.org_id == org_id)).scalars().first()
    if not project:
        return None

    if _is_org_admin(db, user):
        return project

    membership = _project_membership(db, project_id=project_id, user_id=user.id)
    if not membership:
        return None

    if require_edit and not membership.can_edit:
        raise PlannerAuthorizationError("Project membership is read-only")

    return project


def list_planner_projects(db: Session, *, org_id: str, user: User) -> list[tuple[Project, bool]]:
    if _is_org_admin(db, user):
        projects = list(
            db.execute(select(Project).where(Project.org_id == org_id).order_by(Project.name.asc())).scalars().all()
        )
        return [(project, True) for project in projects]

    memberships = list(
        db.execute(
            select(ProjectMember)
            .join(Project, ProjectMember.project_id == Project.id)
            .where(Project.org_id == org_id, ProjectMember.user_id == user.id)
            .options(selectinload(ProjectMember.project))
            .order_by(Project.name.asc())
        )
        .scalars()
        .all()
    )
    return [(membership.project, membership.can_edit) for membership in memberships]


def list_assignable_planners(db: Session, *, org_id: str, user: User) -> list[User]:
    if _is_org_admin(db, user):
        users = list(
            db.execute(
                select(User)
                .where(User.org_id == org_id, User.is_active.is_(True))
                .options(selectinload(User.user_roles).selectinload(UserRole.role))
                .order_by(User.username.asc())
            )
            .scalars()
            .all()
        )
        return [candidate for candidate in users if _is_planner(db, candidate)]

    return [user] if _is_planner(db, user) else []


def list_project_catalog_attractions(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    user: User,
) -> list[tuple[Attraction, str]] | None:
    project = _project_for_user(db, org_id=org_id, project_id=project_id, user=user, require_edit=False)
    if not project:
        return None

    rows = db.execute(
        select(Attraction, Dataset.name)
        .join(ProjectDataset, ProjectDataset.dataset_id == Attraction.dataset_id)
        .join(Dataset, Dataset.id == Attraction.dataset_id)
        .where(
            ProjectDataset.project_id == project_id,
            Attraction.org_id == org_id,
            Attraction.status == "active",
            Attraction.merged_into_attraction_id.is_(None),
        )
        .order_by(Dataset.name.asc(), Attraction.name.asc())
    ).all()
    return [(row[0], row[1]) for row in rows]


def list_project_itineraries(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    user: User,
) -> list[Itinerary] | None:
    project = _project_for_user(db, org_id=org_id, project_id=project_id, user=user, require_edit=False)
    if not project:
        return None

    return list(
        db.execute(
            select(Itinerary)
            .where(Itinerary.org_id == org_id, Itinerary.project_id == project_id)
            .options(selectinload(Itinerary.assigned_planner), selectinload(Itinerary.days))
            .order_by(Itinerary.updated_at.desc())
        )
        .scalars()
        .all()
    )


def _validate_assigned_planner(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    acting_user: User,
    assigned_planner_user_id: str | None,
) -> User | None:
    if not assigned_planner_user_id:
        return None

    assigned = (
        db.execute(
            select(User)
            .where(User.id == assigned_planner_user_id, User.org_id == org_id, User.is_active.is_(True))
            .options(selectinload(User.user_roles).selectinload(UserRole.role))
        )
        .scalars()
        .first()
    )
    if not assigned:
        raise PlannerValidationError("Assigned planner user was not found in this organization")
    if not _is_planner(db, assigned):
        raise PlannerValidationError("Assigned user must have PLANNER role")

    project_member = _project_membership(db, project_id=project_id, user_id=assigned.id)
    if not project_member:
        raise PlannerValidationError("Assigned planner must be a member of the project")

    if not _is_org_admin(db, acting_user) and assigned.id != acting_user.id:
        raise PlannerAuthorizationError("Only ORG_ADMIN can assign other planners")

    return assigned


def _itinerary_with_graph(db: Session, *, itinerary_id: str) -> Itinerary | None:
    return (
        db.execute(
            select(Itinerary)
            .execution_options(populate_existing=True)
            .where(Itinerary.id == itinerary_id)
            .options(
                selectinload(Itinerary.assigned_planner),
                selectinload(Itinerary.days)
                .selectinload(ItineraryDay.stops)
                .selectinload(ItineraryStop.attraction),
            )
        )
        .scalars()
        .first()
    )


def get_itinerary_for_user(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    itinerary_id: str,
    user: User,
    require_edit: bool,
) -> Itinerary | None:
    itinerary = _itinerary_with_graph(db, itinerary_id=itinerary_id)
    if not itinerary:
        return None
    if itinerary.org_id != org_id or itinerary.project_id != project_id:
        return None

    project = _project_for_user(
        db,
        org_id=org_id,
        project_id=project_id,
        user=user,
        require_edit=False,
    )
    if not project:
        return None

    if require_edit:
        if _is_org_admin(db, user):
            return itinerary
        membership = _project_membership(db, project_id=project_id, user_id=user.id)
        if not membership or not membership.can_edit:
            raise PlannerAuthorizationError("Project membership is read-only")
        if itinerary.assigned_planner_user_id and itinerary.assigned_planner_user_id != user.id:
            raise PlannerAuthorizationError("Itinerary is assigned to another planner")

    return itinerary


def _haversine_miles(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    lat1_rad = math.radians(lat1)
    lat2_rad = math.radians(lat2)
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)

    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(lat1_rad) * math.cos(lat2_rad) * (math.sin(dlon / 2) ** 2)
    )
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return EARTH_RADIUS_MILES * c


def _format_clock(minute_of_day: int) -> str:
    hour = minute_of_day // 60
    minute = minute_of_day % 60
    return f"{hour:02d}:{minute:02d}"


def _day_effective_speeds(org: Organization, itinerary: Itinerary, day: ItineraryDay) -> tuple[float, float]:
    urban = (
        day.urban_speed_mph_override
        if day.urban_speed_mph_override is not None
        else itinerary.urban_speed_mph_override
        if itinerary.urban_speed_mph_override is not None
        else org.default_urban_speed_mph
    )
    highway = (
        day.highway_speed_mph_override
        if day.highway_speed_mph_override is not None
        else itinerary.highway_speed_mph_override
        if itinerary.highway_speed_mph_override is not None
        else org.default_highway_speed_mph
    )
    return urban, highway


def analyze_day(org: Organization, itinerary: Itinerary, day: ItineraryDay) -> dict:
    ordered_stops = sorted(day.stops, key=lambda stop: stop.order_index)
    effective_urban, effective_highway = _day_effective_speeds(org, itinerary, day)

    travel_distance = 0.0
    travel_minutes = 0.0
    for idx in range(len(ordered_stops) - 1):
        current_stop = ordered_stops[idx]
        next_stop = ordered_stops[idx + 1]
        distance = _haversine_miles(
            current_stop.attraction.latitude,
            current_stop.attraction.longitude,
            next_stop.attraction.latitude,
            next_stop.attraction.longitude,
        )
        travel_distance += distance

        urban_distance = min(distance, URBAN_DISTANCE_PORTION_MILES)
        highway_distance = max(distance - URBAN_DISTANCE_PORTION_MILES, 0.0)
        travel_minutes += (urban_distance / effective_urban) * 60.0
        travel_minutes += (highway_distance / effective_highway) * 60.0

    warnings: list[dict[str, str]] = []
    for idx in range(len(ordered_stops) - 1):
        current_stop = ordered_stops[idx]
        next_stop = ordered_stops[idx + 1]
        overlap_minutes = (current_stop.start_minute_of_day + current_stop.duration_minutes) - next_stop.start_minute_of_day
        if overlap_minutes >= 15:
            warnings.append(
                {
                    "code": "overlap_15m",
                    "message": (
                        f"{current_stop.attraction.name} overlaps {next_stop.attraction.name} by "
                        f"{overlap_minutes} minutes ({_format_clock(current_stop.start_minute_of_day)}-"
                        f"{_format_clock(current_stop.start_minute_of_day + current_stop.duration_minutes)} vs "
                        f"{_format_clock(next_stop.start_minute_of_day)})."
                    ),
                }
            )

    activity_minutes = sum(stop.duration_minutes for stop in ordered_stops)
    if activity_minutes > 12 * 60:
        warnings.append(
            {
                "code": "activity_exceeds_12h",
                "message": f"Activity time is {activity_minutes} minutes, above 12 hours.",
            }
        )

    return {
        "effective_urban_speed_mph": round(effective_urban, 2),
        "effective_highway_speed_mph": round(effective_highway, 2),
        "travel_distance_miles": round(travel_distance, 2),
        "travel_time_minutes": int(round(travel_minutes)),
        "activity_minutes": activity_minutes,
        "warnings": warnings,
        "ordered_stops": ordered_stops,
    }


def _serialize_snapshot(itinerary: Itinerary, org: Organization) -> dict:
    snapshot_days = []
    for day in sorted(itinerary.days, key=lambda row: row.day_number):
        day_analysis = analyze_day(org, itinerary, day)
        snapshot_days.append(
            {
                "id": day.id,
                "day_number": day.day_number,
                "title": day.title,
                "notes": day.notes,
                "urban_speed_mph_override": day.urban_speed_mph_override,
                "highway_speed_mph_override": day.highway_speed_mph_override,
                "effective_urban_speed_mph": day_analysis["effective_urban_speed_mph"],
                "effective_highway_speed_mph": day_analysis["effective_highway_speed_mph"],
                "travel_distance_miles": day_analysis["travel_distance_miles"],
                "travel_time_minutes": day_analysis["travel_time_minutes"],
                "activity_minutes": day_analysis["activity_minutes"],
                "warnings": day_analysis["warnings"],
                "stops": [
                    {
                        "id": stop.id,
                        "attraction_id": stop.attraction_id,
                        "attraction_name": stop.attraction.name,
                        "order_index": stop.order_index,
                        "start_minute_of_day": stop.start_minute_of_day,
                        "duration_minutes": stop.duration_minutes,
                        "notes": stop.notes,
                    }
                    for stop in day_analysis["ordered_stops"]
                ],
            }
        )

    return {
        "itinerary": {
            "id": itinerary.id,
            "org_id": itinerary.org_id,
            "project_id": itinerary.project_id,
            "name": itinerary.name,
            "description": itinerary.description,
            "status": itinerary.status,
            "assigned_planner_user_id": itinerary.assigned_planner_user_id,
            "urban_speed_mph_override": itinerary.urban_speed_mph_override,
            "highway_speed_mph_override": itinerary.highway_speed_mph_override,
            "org_default_urban_speed_mph": org.default_urban_speed_mph,
            "org_default_highway_speed_mph": org.default_highway_speed_mph,
            "created_at": itinerary.created_at.astimezone(UTC).isoformat(),
            "updated_at": itinerary.updated_at.astimezone(UTC).isoformat(),
        },
        "days": snapshot_days,
    }


def _record_version(db: Session, *, itinerary_id: str, changed_by_user_id: str, change_summary: str) -> ItineraryVersion:
    itinerary = _itinerary_with_graph(db, itinerary_id=itinerary_id)
    if not itinerary:
        raise PlannerValidationError("Itinerary was not found for version snapshot")

    locked_itinerary = db.execute(
        select(Itinerary)
        .where(Itinerary.id == itinerary_id)
        .execution_options(populate_existing=True)
        .with_for_update()
    ).scalars().one()

    org = db.execute(select(Organization).where(Organization.id == itinerary.org_id)).scalars().one()
    next_version_number = locked_itinerary.version_counter + 1
    locked_itinerary.version_counter = next_version_number
    db.flush()

    version = ItineraryVersion(
        org_id=itinerary.org_id,
        project_id=itinerary.project_id,
        itinerary_id=itinerary.id,
        version_number=next_version_number,
        change_summary=change_summary,
        snapshot=_serialize_snapshot(itinerary, org),
        created_by_user_id=changed_by_user_id,
    )
    db.add(locked_itinerary)
    db.add(version)
    db.commit()
    db.refresh(version)
    return version


def _normalize_column_name(name: str) -> str:
    return re.sub(r"\s+", "_", name.strip().lower())


def _clock_to_minutes(clock_text: str) -> int | None:
    match = re.fullmatch(r"(\d{1,2}):(\d{2})", clock_text.strip())
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2))
    if hour < 0 or hour > 23 or minute < 0 or minute > 59:
        return None
    return hour * 60 + minute


def _project_catalog_map(db: Session, *, org_id: str, project_id: str) -> dict[str, Attraction]:
    rows = db.execute(
        select(Attraction)
        .join(ProjectDataset, ProjectDataset.dataset_id == Attraction.dataset_id)
        .where(
            ProjectDataset.project_id == project_id,
            Attraction.org_id == org_id,
            Attraction.status == "active",
            Attraction.merged_into_attraction_id.is_(None),
        )
    ).scalars()
    return {row.id: row for row in rows}


def _coerce_xlsx_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _extract_format(file_name: str) -> str | None:
    lowered = file_name.lower()
    if lowered.endswith(".csv"):
        return "csv"
    if lowered.endswith(".xlsx"):
        return "xlsx"
    return None


def _inspect_zip_payload(
    payload: bytes,
    *,
    max_entries: int,
    max_uncompressed_bytes: int,
) -> tuple[list[zipfile.ZipInfo], set[str]]:
    try:
        with zipfile.ZipFile(io.BytesIO(payload), mode="r") as archive:
            infos = [info for info in archive.infolist() if not info.is_dir()]
    except zipfile.BadZipFile as exc:
        raise PlannerValidationError("Uploaded ZIP content is invalid.") from exc

    if len(infos) > max_entries:
        raise PlannerValidationError(f"Uploaded ZIP content exceeds the maximum of {max_entries} files.")

    total_uncompressed_bytes = 0
    for info in infos:
        total_uncompressed_bytes += info.file_size
        if total_uncompressed_bytes > max_uncompressed_bytes:
            raise PlannerValidationError(
                f"Uploaded ZIP content exceeds the maximum uncompressed size of {max_uncompressed_bytes // (1024 * 1024)} MB."
            )

    return infos, {info.filename for info in infos}


def _detect_csv_import_mime(payload: bytes) -> str | None:
    if b"\x00" in payload:
        return None

    try:
        decoded = payload.decode("utf-8-sig")
    except UnicodeDecodeError:
        return None

    if not decoded.strip():
        return None

    sample = decoded[:8192]
    if any((ord(ch) < 32 and ch not in {"\n", "\r", "\t"}) for ch in sample):
        return None

    try:
        dialect = csv.Sniffer().sniff(sample)
        if dialect.delimiter not in {",", ";", "\t", "|"}:
            return None
    except csv.Error:
        if "," not in sample and "\n" not in sample:
            return None

    return "text/csv"


def _detect_xlsx_import_mime(payload: bytes) -> str | None:
    settings = get_settings()
    _, names = _inspect_zip_payload(
        payload,
        max_entries=settings.planner_import_archive_max_entries,
        max_uncompressed_bytes=settings.planner_import_archive_max_uncompressed_bytes,
    )
    if "[Content_Types].xml" not in names:
        return None
    if not any(name.startswith("xl/") for name in names):
        return None
    return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _validate_itinerary_import_payload(file_name: str, content: bytes) -> str:
    file_format = _extract_format(file_name)
    if file_format is None:
        raise PlannerValidationError("Unsupported file type. Upload a .csv or .xlsx file.")

    if file_format == "csv":
        detected_mime = _detect_csv_import_mime(content)
        if detected_mime != "text/csv":
            raise PlannerValidationError("Uploaded .csv file content does not match CSV format.")
        return file_format

    detected_mime = _detect_xlsx_import_mime(content)
    if detected_mime != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        raise PlannerValidationError("Uploaded .xlsx file content does not match XLSX format.")
    return file_format


def _parse_csv_rows(content: bytes) -> tuple[list[dict[str, str]], list[str]]:
    errors: list[str] = []
    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        errors.append("CSV file must be UTF-8 encoded.")
        return [], errors

    reader = csv.DictReader(io.StringIO(decoded))
    if not reader.fieldnames:
        errors.append("CSV header row is missing.")
        return [], errors

    normalized_to_actual = {_normalize_column_name(name): name for name in reader.fieldnames if name is not None}
    missing = [column for column in REQUIRED_IMPORT_COLUMNS if column not in normalized_to_actual]
    if missing:
        errors.append(
            "Missing required columns: " + ", ".join(missing) + ". Download an export file and reuse its headers."
        )
        return [], errors

    rows: list[dict[str, str]] = []
    for row in reader:
        normalized_row = {column: (row.get(normalized_to_actual.get(column, "")) or "").strip() for column in IMPORT_EXPORT_COLUMNS}
        if not any(normalized_row.values()):
            continue
        rows.append(normalized_row)
    return rows, errors


def _parse_xlsx_rows(content: bytes) -> tuple[list[dict[str, str]], list[str]]:
    errors: list[str] = []
    if load_workbook is None:
        errors.append("XLSX support is unavailable in this runtime.")
        return [], errors

    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        errors.append("XLSX file could not be opened. Ensure it is a valid .xlsx workbook.")
        return [], errors

    sheet = workbook.active
    raw_headers = [_coerce_xlsx_cell(cell.value) for cell in sheet[1]] if sheet.max_row >= 1 else []
    if not raw_headers:
        errors.append("XLSX header row is missing.")
        return [], errors

    normalized_to_index: dict[str, int] = {}
    for index, header in enumerate(raw_headers):
        normalized = _normalize_column_name(header)
        if normalized:
            normalized_to_index[normalized] = index

    missing = [column for column in REQUIRED_IMPORT_COLUMNS if column not in normalized_to_index]
    if missing:
        errors.append(
            "Missing required columns: " + ", ".join(missing) + ". Download an export file and reuse its headers."
        )
        return [], errors

    rows: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=False):
        normalized_row: dict[str, str] = {}
        for column in IMPORT_EXPORT_COLUMNS:
            index = normalized_to_index.get(column)
            normalized_row[column] = _coerce_xlsx_cell(row[index].value) if index is not None and index < len(row) else ""
        if not any(normalized_row.values()):
            continue
        rows.append(normalized_row)
    return rows, errors


def _export_rows_for_itinerary(itinerary: Itinerary) -> list[dict[str, str | int | float]]:
    rows: list[dict[str, str | int | float]] = []
    for day in sorted(itinerary.days, key=lambda row: row.day_number):
        for stop in sorted(day.stops, key=lambda row: row.order_index):
            rows.append(
                {
                    "day_number": day.day_number,
                    "day_title": day.title or "",
                    "day_notes": day.notes or "",
                    "day_urban_speed_mph_override": day.urban_speed_mph_override or "",
                    "day_highway_speed_mph_override": day.highway_speed_mph_override or "",
                    "stop_order": stop.order_index + 1,
                    "attraction_id": stop.attraction_id,
                    "attraction_name": stop.attraction.name,
                    "attraction_city": stop.attraction.city,
                    "attraction_state": stop.attraction.state,
                    "start_time": _format_clock(stop.start_minute_of_day),
                    "duration_minutes": stop.duration_minutes,
                    "stop_notes": stop.notes or "",
                }
            )
    return rows


def export_itinerary_file(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    itinerary_id: str,
    user: User,
    export_format: str,
) -> tuple[bytes, str, str, str] | None:
    itinerary = get_itinerary_for_user(
        db,
        org_id=org_id,
        project_id=project_id,
        itinerary_id=itinerary_id,
        user=user,
        require_edit=False,
    )
    if not itinerary:
        return None

    normalized_format = export_format.lower()
    if normalized_format not in SUPPORTED_IMPORT_FORMATS:
        raise PlannerValidationError("Export format must be csv or xlsx")

    rows = _export_rows_for_itinerary(itinerary)
    if normalized_format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=IMPORT_EXPORT_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
        payload = output.getvalue().encode("utf-8")
        return payload, "text/csv; charset=utf-8", "csv", itinerary.name

    if Workbook is None:
        raise PlannerValidationError("XLSX export support is unavailable in this runtime")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "itinerary"
    sheet.append(IMPORT_EXPORT_COLUMNS)
    for row in rows:
        sheet.append([row[column] for column in IMPORT_EXPORT_COLUMNS])

    binary = io.BytesIO()
    workbook.save(binary)
    return binary.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx", itinerary.name


def _build_rejection_index() -> dict[int, dict]:
    return {}


def _reject_row(
    rejected_index: dict[int, dict],
    *,
    row_number: int,
    raw_row: dict[str, str],
    error: str,
    hint: str,
) -> None:
    row = rejected_index.setdefault(
        row_number,
        {
            "row_number": row_number,
            "raw_row": raw_row,
            "errors": [],
            "correction_hints": [],
        },
    )
    if error not in row["errors"]:
        row["errors"].append(error)
    if hint and hint not in row["correction_hints"]:
        row["correction_hints"].append(hint)


def import_itinerary_file(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    itinerary_id: str,
    user: User,
    file_name: str,
    content: bytes,
) -> dict | None:
    itinerary = get_itinerary_for_user(
        db,
        org_id=org_id,
        project_id=project_id,
        itinerary_id=itinerary_id,
        user=user,
        require_edit=True,
    )
    if not itinerary:
        return None

    file_format = _validate_itinerary_import_payload(file_name, content)
    file_errors: list[str] = []
    if file_format == "csv":
        parsed_rows, parse_errors = _parse_csv_rows(content)
        file_errors.extend(parse_errors)
    else:
        parsed_rows, parse_errors = _parse_xlsx_rows(content)
        file_errors.extend(parse_errors)

    catalog_map = _project_catalog_map(db, org_id=org_id, project_id=project_id)
    attraction_name_key_map = {
        (row.name.strip().lower(), row.city.strip().lower(), row.state.strip().lower()): row for row in catalog_map.values()
    }

    rejected_index = _build_rejection_index()
    accepted_candidates: list[dict] = []

    for offset, row in enumerate(parsed_rows):
        row_number = offset + 2
        errors: list[tuple[str, str]] = []

        day_number_raw = row["day_number"].strip()
        if not day_number_raw:
            errors.append(("day_number is required.", "Provide a day_number between 1 and 365."))
            day_number = None
        else:
            try:
                day_number = int(day_number_raw)
                if day_number < 1 or day_number > 365:
                    errors.append(("day_number must be between 1 and 365.", "Use an integer day_number within 1..365."))
            except ValueError:
                day_number = None
                errors.append(("day_number must be an integer.", "Use numeric day_number values such as 1, 2, 3."))

        stop_order_raw = row["stop_order"].strip()
        if not stop_order_raw:
            errors.append(("stop_order is required.", "Provide stop_order as a positive integer starting at 1."))
            stop_order = None
        else:
            try:
                stop_order = int(stop_order_raw)
                if stop_order < 1:
                    errors.append(("stop_order must be at least 1.", "Renumber stop_order values to start at 1."))
            except ValueError:
                stop_order = None
                errors.append(("stop_order must be an integer.", "Use numeric stop_order values such as 1, 2, 3."))

        attraction_id = row["attraction_id"].strip()
        attraction = catalog_map.get(attraction_id)
        if not attraction_id:
            errors.append(("attraction_id is required.", "Use attraction_id values from itinerary export or project catalog."))
        elif not attraction:
            hint = "Ensure attraction_id belongs to an active attraction linked to this project."
            if row["attraction_name"].strip():
                name_key = (
                    row["attraction_name"].strip().lower(),
                    row["attraction_city"].strip().lower(),
                    row["attraction_state"].strip().lower(),
                )
                guessed = attraction_name_key_map.get(name_key)
                if guessed:
                    hint = f"Use attraction_id {guessed.id} for {guessed.name}."
            errors.append(("attraction_id is not available in this project's active catalog.", hint))

        start_time = row["start_time"].strip()
        start_minute = _clock_to_minutes(start_time)
        if not start_time:
            errors.append(("start_time is required.", "Provide start_time in 24-hour HH:MM format."))
        elif start_minute is None:
            errors.append(("start_time must be HH:MM in 24-hour time.", "Use values like 09:30 or 14:05."))

        duration_raw = row["duration_minutes"].strip()
        if not duration_raw:
            errors.append(("duration_minutes is required.", "Provide duration_minutes between 5 and 720."))
            duration_minutes = None
        else:
            try:
                duration_minutes = int(duration_raw)
                if duration_minutes < 5 or duration_minutes > 720:
                    errors.append(
                        ("duration_minutes must be between 5 and 720.", "Choose a duration within the 5..720 range.")
                    )
            except ValueError:
                duration_minutes = None
                errors.append(("duration_minutes must be an integer.", "Use whole-minute values such as 45 or 120."))

        def parse_optional_speed(field_name: str) -> tuple[float | None, tuple[str, str] | None]:
            value = row[field_name].strip()
            if not value:
                return None, None
            try:
                parsed = float(value)
            except ValueError:
                return None, (f"{field_name} must be numeric when provided.", "Use a positive numeric speed value.")
            if parsed <= 0:
                return None, (f"{field_name} must be greater than 0.", "Use a positive numeric speed value.")
            return parsed, None

        day_urban_speed, day_urban_error = parse_optional_speed("day_urban_speed_mph_override")
        day_highway_speed, day_highway_error = parse_optional_speed("day_highway_speed_mph_override")
        if day_urban_error:
            errors.append(day_urban_error)
        if day_highway_error:
            errors.append(day_highway_error)

        if errors:
            for error_message, hint in errors:
                _reject_row(
                    rejected_index,
                    row_number=row_number,
                    raw_row=row,
                    error=error_message,
                    hint=hint,
                )
            continue

        accepted_candidates.append(
            {
                "row_number": row_number,
                "raw_row": row,
                "day_number": day_number,
                "day_title": row["day_title"].strip() or None,
                "day_notes": row["day_notes"].strip() or None,
                "day_urban_speed_mph_override": day_urban_speed,
                "day_highway_speed_mph_override": day_highway_speed,
                "stop_order": stop_order,
                "attraction_id": attraction_id,
                "attraction_name": attraction.name if attraction else "",
                "start_time": f"{start_minute // 60:02d}:{start_minute % 60:02d}",
                "start_minute_of_day": start_minute,
                "duration_minutes": duration_minutes,
                "stop_notes": row["stop_notes"].strip() or None,
            }
        )

    key_to_first_row: dict[tuple[int, int], int] = {}
    for row in accepted_candidates:
        key = (row["day_number"], row["stop_order"])
        if key in key_to_first_row:
            first_row_number = key_to_first_row[key]
            _reject_row(
                rejected_index,
                row_number=first_row_number,
                raw_row=next(candidate["raw_row"] for candidate in accepted_candidates if candidate["row_number"] == first_row_number),
                error="Duplicate stop_order detected for the same day_number.",
                hint="Ensure each day uses unique stop_order values.",
            )
            _reject_row(
                rejected_index,
                row_number=row["row_number"],
                raw_row=row["raw_row"],
                error="Duplicate stop_order detected for the same day_number.",
                hint="Ensure each day uses unique stop_order values.",
            )
        else:
            key_to_first_row[key] = row["row_number"]

    rows_by_day: dict[int, list[dict]] = {}
    for row in accepted_candidates:
        rows_by_day.setdefault(row["day_number"], []).append(row)

    for day_number, day_rows in rows_by_day.items():
        baseline = day_rows[0]
        baseline_signature = (
            baseline["day_title"],
            baseline["day_notes"],
            baseline["day_urban_speed_mph_override"],
            baseline["day_highway_speed_mph_override"],
        )
        for row in day_rows[1:]:
            row_signature = (
                row["day_title"],
                row["day_notes"],
                row["day_urban_speed_mph_override"],
                row["day_highway_speed_mph_override"],
            )
            if row_signature != baseline_signature:
                _reject_row(
                    rejected_index,
                    row_number=row["row_number"],
                    raw_row=row["raw_row"],
                    error=f"Day metadata conflicts for day_number {day_number}.",
                    hint="Use identical day_title/day_notes/day speed override values on all rows for the same day.",
                )

    accepted_rows = [row for row in accepted_candidates if row["row_number"] not in rejected_index]

    rows_by_day_after_filter: dict[int, list[dict]] = {}
    for row in accepted_rows:
        rows_by_day_after_filter.setdefault(row["day_number"], []).append(row)

    for day_number, day_rows in rows_by_day_after_filter.items():
        ordered = sorted(day_rows, key=lambda item: item["stop_order"])
        expected = list(range(1, len(ordered) + 1))
        actual = [item["stop_order"] for item in ordered]
        if actual != expected:
            for row in ordered:
                _reject_row(
                    rejected_index,
                    row_number=row["row_number"],
                    raw_row=row["raw_row"],
                    error=f"stop_order values for day_number {day_number} must be contiguous starting at 1.",
                    hint="Renumber stop_order to 1..N within each day.",
                )

    accepted_rows = [row for row in accepted_rows if row["row_number"] not in rejected_index]

    applied = False
    applied_day_count = 0
    applied_stop_count = 0

    if not file_errors and accepted_rows:
        for existing_day in list(itinerary.days):
            db.delete(existing_day)
        db.flush()

        for day_number in sorted({row["day_number"] for row in accepted_rows}):
            day_rows = [row for row in accepted_rows if row["day_number"] == day_number]
            day_rows.sort(key=lambda item: item["stop_order"])
            day = ItineraryDay(
                itinerary_id=itinerary.id,
                day_number=day_number,
                title=day_rows[0]["day_title"],
                notes=day_rows[0]["day_notes"],
                urban_speed_mph_override=day_rows[0]["day_urban_speed_mph_override"],
                highway_speed_mph_override=day_rows[0]["day_highway_speed_mph_override"],
            )
            db.add(day)
            db.flush()
            applied_day_count += 1

            for row in day_rows:
                stop = ItineraryStop(
                    itinerary_day_id=day.id,
                    attraction_id=row["attraction_id"],
                    order_index=row["stop_order"] - 1,
                    start_minute_of_day=row["start_minute_of_day"],
                    duration_minutes=row["duration_minutes"],
                    notes=row["stop_notes"],
                )
                db.add(stop)
                applied_stop_count += 1

        itinerary.updated_by_user_id = user.id
        db.add(itinerary)
        db.flush()
        _record_version(
            db,
            itinerary_id=itinerary.id,
            changed_by_user_id=user.id,
            change_summary=(
                f"Itinerary import applied ({len(accepted_rows)} accepted, {len(rejected_index)} rejected)"
            ),
        )
        applied = True

    rejected_rows = sorted(rejected_index.values(), key=lambda row: row["row_number"])
    accepted_payload = [
        {
            "row_number": row["row_number"],
            "day_number": row["day_number"],
            "stop_order": row["stop_order"],
            "attraction_id": row["attraction_id"],
            "attraction_name": row["attraction_name"],
            "start_time": row["start_time"],
            "duration_minutes": row["duration_minutes"],
        }
        for row in sorted(accepted_rows, key=lambda item: item["row_number"])
    ]

    return {
        "itinerary_id": itinerary.id,
        "project_id": itinerary.project_id,
        "file_name": file_name,
        "file_format": file_format or "unknown",
        "imported_at": datetime.now(UTC),
        "applied": applied,
        "total_rows": len(parsed_rows),
        "accepted_row_count": len(accepted_payload),
        "rejected_row_count": len(rejected_rows),
        "applied_day_count": applied_day_count,
        "applied_stop_count": applied_stop_count,
        "file_errors": file_errors,
        "accepted_rows": accepted_payload,
        "rejected_rows": rejected_rows,
    }


def create_itinerary(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    user: User,
    name: str,
    description: str | None,
    status: str,
    assigned_planner_user_id: str | None,
    urban_speed_mph_override: float | None,
    highway_speed_mph_override: float | None,
) -> Itinerary | None:
    project = _project_for_user(db, org_id=org_id, project_id=project_id, user=user, require_edit=True)
    if not project:
        return None

    _validate_assigned_planner(
        db,
        org_id=org_id,
        project_id=project_id,
        acting_user=user,
        assigned_planner_user_id=assigned_planner_user_id,
    )

    itinerary = Itinerary(
        org_id=org_id,
        project_id=project_id,
        name=name.strip(),
        description=description,
        status=status.strip(),
        assigned_planner_user_id=assigned_planner_user_id,
        urban_speed_mph_override=urban_speed_mph_override,
        highway_speed_mph_override=highway_speed_mph_override,
        created_by_user_id=user.id,
        updated_by_user_id=user.id,
    )
    db.add(itinerary)
    try:
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise PlannerConflictError("Itinerary name must be unique within project") from exc

    _record_version(db, itinerary_id=itinerary.id, changed_by_user_id=user.id, change_summary="Itinerary created")
    return _itinerary_with_graph(db, itinerary_id=itinerary.id)


def update_itinerary(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    itinerary_id: str,
    user: User,
    name: str | None,
    description: str | None,
    status: str | None,
    assigned_planner_user_id: str | None,
    urban_speed_mph_override: float | None,
    highway_speed_mph_override: float | None,
    provided_fields: set[str],
) -> Itinerary | None:
    itinerary = get_itinerary_for_user(
        db,
        org_id=org_id,
        project_id=project_id,
        itinerary_id=itinerary_id,
        user=user,
        require_edit=True,
    )
    if not itinerary:
        return None

    if "name" in provided_fields:
        if name is None:
            raise PlannerValidationError("Itinerary name cannot be null")
        itinerary.name = name.strip()
    if "description" in provided_fields:
        itinerary.description = description
    if "status" in provided_fields:
        if status is None:
            raise PlannerValidationError("Itinerary status cannot be null")
        itinerary.status = status.strip()
    if "urban_speed_mph_override" in provided_fields:
        itinerary.urban_speed_mph_override = urban_speed_mph_override
    if "highway_speed_mph_override" in provided_fields:
        itinerary.highway_speed_mph_override = highway_speed_mph_override

    if "assigned_planner_user_id" in provided_fields:
        _validate_assigned_planner(
            db,
            org_id=org_id,
            project_id=project_id,
            acting_user=user,
            assigned_planner_user_id=assigned_planner_user_id,
        )
        itinerary.assigned_planner_user_id = assigned_planner_user_id

    itinerary.updated_by_user_id = user.id
    db.add(itinerary)
    try:
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise PlannerConflictError("Itinerary name must be unique within project") from exc

    _record_version(db, itinerary_id=itinerary.id, changed_by_user_id=user.id, change_summary="Itinerary updated")
    return _itinerary_with_graph(db, itinerary_id=itinerary.id)


def archive_itinerary(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    itinerary_id: str,
    user: User,
) -> bool:
    itinerary = get_itinerary_for_user(
        db,
        org_id=org_id,
        project_id=project_id,
        itinerary_id=itinerary_id,
        user=user,
        require_edit=True,
    )
    if not itinerary:
        return False

    itinerary.status = "archived"
    itinerary.updated_by_user_id = user.id
    db.add(itinerary)
    db.flush()
    _record_version(db, itinerary_id=itinerary.id, changed_by_user_id=user.id, change_summary="Itinerary archived")
    return True


def create_itinerary_day(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    itinerary_id: str,
    user: User,
    day_number: int,
    title: str | None,
    notes: str | None,
    urban_speed_mph_override: float | None,
    highway_speed_mph_override: float | None,
) -> Itinerary | None:
    itinerary = get_itinerary_for_user(
        db,
        org_id=org_id,
        project_id=project_id,
        itinerary_id=itinerary_id,
        user=user,
        require_edit=True,
    )
    if not itinerary:
        return None

    day = ItineraryDay(
        itinerary_id=itinerary.id,
        day_number=day_number,
        title=title,
        notes=notes,
        urban_speed_mph_override=urban_speed_mph_override,
        highway_speed_mph_override=highway_speed_mph_override,
    )
    db.add(day)
    try:
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise PlannerConflictError("Day number must be unique within itinerary") from exc

    itinerary.updated_by_user_id = user.id
    db.add(itinerary)
    db.flush()
    _record_version(db, itinerary_id=itinerary.id, changed_by_user_id=user.id, change_summary="Itinerary day added")
    return _itinerary_with_graph(db, itinerary_id=itinerary.id)


def update_itinerary_day(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    itinerary_id: str,
    day_id: str,
    user: User,
    day_number: int | None,
    title: str | None,
    notes: str | None,
    urban_speed_mph_override: float | None,
    highway_speed_mph_override: float | None,
    provided_fields: set[str],
) -> Itinerary | None:
    itinerary = get_itinerary_for_user(
        db,
        org_id=org_id,
        project_id=project_id,
        itinerary_id=itinerary_id,
        user=user,
        require_edit=True,
    )
    if not itinerary:
        return None

    day = next((row for row in itinerary.days if row.id == day_id), None)
    if not day:
        return None

    if "day_number" in provided_fields:
        if day_number is None:
            raise PlannerValidationError("Day number cannot be null")
        day.day_number = day_number
    if "title" in provided_fields:
        day.title = title
    if "notes" in provided_fields:
        day.notes = notes
    if "urban_speed_mph_override" in provided_fields:
        day.urban_speed_mph_override = urban_speed_mph_override
    if "highway_speed_mph_override" in provided_fields:
        day.highway_speed_mph_override = highway_speed_mph_override

    itinerary.updated_by_user_id = user.id
    db.add(day)
    db.add(itinerary)
    try:
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise PlannerConflictError("Day number must be unique within itinerary") from exc

    _record_version(db, itinerary_id=itinerary.id, changed_by_user_id=user.id, change_summary="Itinerary day updated")
    return _itinerary_with_graph(db, itinerary_id=itinerary.id)


def delete_itinerary_day(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    itinerary_id: str,
    day_id: str,
    user: User,
) -> Itinerary | None:
    itinerary = get_itinerary_for_user(
        db,
        org_id=org_id,
        project_id=project_id,
        itinerary_id=itinerary_id,
        user=user,
        require_edit=True,
    )
    if not itinerary:
        return None

    day = next((row for row in itinerary.days if row.id == day_id), None)
    if not day:
        return None

    db.delete(day)
    itinerary.updated_by_user_id = user.id
    db.add(itinerary)
    db.flush()
    _record_version(db, itinerary_id=itinerary.id, changed_by_user_id=user.id, change_summary="Itinerary day removed")
    return _itinerary_with_graph(db, itinerary_id=itinerary.id)


def _attraction_in_project_catalog(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    attraction_id: str,
) -> Attraction | None:
    return (
        db.execute(
            select(Attraction)
            .join(ProjectDataset, ProjectDataset.dataset_id == Attraction.dataset_id)
            .where(
                Attraction.id == attraction_id,
                Attraction.org_id == org_id,
                ProjectDataset.project_id == project_id,
                Attraction.merged_into_attraction_id.is_(None),
            )
        )
        .scalars()
        .first()
    )


def create_itinerary_stop(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    itinerary_id: str,
    day_id: str,
    user: User,
    attraction_id: str,
    start_minute_of_day: int,
    duration_minutes: int,
    notes: str | None,
) -> Itinerary | None:
    itinerary = get_itinerary_for_user(
        db,
        org_id=org_id,
        project_id=project_id,
        itinerary_id=itinerary_id,
        user=user,
        require_edit=True,
    )
    if not itinerary:
        return None

    day = next((row for row in itinerary.days if row.id == day_id), None)
    if not day:
        return None

    attraction = _attraction_in_project_catalog(
        db,
        org_id=org_id,
        project_id=project_id,
        attraction_id=attraction_id,
    )
    if not attraction:
        raise PlannerValidationError("Attraction is not available through this project's linked catalog")

    next_order_index = (max((stop.order_index for stop in day.stops), default=-1) + 1) if day.stops else 0
    stop = ItineraryStop(
        itinerary_day_id=day.id,
        attraction_id=attraction.id,
        order_index=next_order_index,
        start_minute_of_day=start_minute_of_day,
        duration_minutes=duration_minutes,
        notes=notes,
    )
    db.add(stop)
    itinerary.updated_by_user_id = user.id
    db.add(itinerary)
    db.flush()
    _record_version(db, itinerary_id=itinerary.id, changed_by_user_id=user.id, change_summary="Itinerary stop added")
    return _itinerary_with_graph(db, itinerary_id=itinerary.id)


def update_itinerary_stop(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    itinerary_id: str,
    day_id: str,
    stop_id: str,
    user: User,
    start_minute_of_day: int | None,
    duration_minutes: int | None,
    notes: str | None,
    provided_fields: set[str],
) -> Itinerary | None:
    itinerary = get_itinerary_for_user(
        db,
        org_id=org_id,
        project_id=project_id,
        itinerary_id=itinerary_id,
        user=user,
        require_edit=True,
    )
    if not itinerary:
        return None

    day = next((row for row in itinerary.days if row.id == day_id), None)
    if not day:
        return None
    stop = next((row for row in day.stops if row.id == stop_id), None)
    if not stop:
        return None

    if "start_minute_of_day" in provided_fields:
        if start_minute_of_day is None:
            raise PlannerValidationError("Stop start time cannot be null")
        stop.start_minute_of_day = start_minute_of_day
    if "duration_minutes" in provided_fields:
        if duration_minutes is None:
            raise PlannerValidationError("Stop duration cannot be null")
        stop.duration_minutes = duration_minutes
    if "notes" in provided_fields:
        stop.notes = notes

    itinerary.updated_by_user_id = user.id
    db.add(stop)
    db.add(itinerary)
    db.flush()
    _record_version(db, itinerary_id=itinerary.id, changed_by_user_id=user.id, change_summary="Itinerary stop updated")
    return _itinerary_with_graph(db, itinerary_id=itinerary.id)


def delete_itinerary_stop(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    itinerary_id: str,
    day_id: str,
    stop_id: str,
    user: User,
) -> Itinerary | None:
    itinerary = get_itinerary_for_user(
        db,
        org_id=org_id,
        project_id=project_id,
        itinerary_id=itinerary_id,
        user=user,
        require_edit=True,
    )
    if not itinerary:
        return None

    day = next((row for row in itinerary.days if row.id == day_id), None)
    if not day:
        return None
    stop = next((row for row in day.stops if row.id == stop_id), None)
    if not stop:
        return None

    remaining_stops = [row for row in day.stops if row.id != stop_id]
    db.delete(stop)
    db.flush()
    for index, remaining in enumerate(sorted(remaining_stops, key=lambda row: row.order_index)):
        remaining.order_index = index
        db.add(remaining)

    itinerary.updated_by_user_id = user.id
    db.add(itinerary)
    db.flush()
    _record_version(db, itinerary_id=itinerary.id, changed_by_user_id=user.id, change_summary="Itinerary stop removed")
    return _itinerary_with_graph(db, itinerary_id=itinerary.id)


def reorder_itinerary_stops(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    itinerary_id: str,
    day_id: str,
    user: User,
    ordered_stop_ids: list[str],
) -> Itinerary | None:
    itinerary = get_itinerary_for_user(
        db,
        org_id=org_id,
        project_id=project_id,
        itinerary_id=itinerary_id,
        user=user,
        require_edit=True,
    )
    if not itinerary:
        return None

    day = next((row for row in itinerary.days if row.id == day_id), None)
    if not day:
        return None

    stop_by_id = {stop.id: stop for stop in day.stops}
    if set(stop_by_id.keys()) != set(ordered_stop_ids):
        raise PlannerValidationError("Reorder payload must include each stop exactly once")

    offset = len(ordered_stop_ids)
    for index, stop_id in enumerate(ordered_stop_ids):
        stop_by_id[stop_id].order_index = index + offset
        db.add(stop_by_id[stop_id])
    db.flush()

    for index, stop_id in enumerate(ordered_stop_ids):
        stop_by_id[stop_id].order_index = index
        db.add(stop_by_id[stop_id])

    itinerary.updated_by_user_id = user.id
    db.add(itinerary)
    db.flush()
    _record_version(db, itinerary_id=itinerary.id, changed_by_user_id=user.id, change_summary="Itinerary stops reordered")
    return _itinerary_with_graph(db, itinerary_id=itinerary.id)


def list_itinerary_versions(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    itinerary_id: str,
    user: User,
) -> list[ItineraryVersion] | None:
    itinerary = get_itinerary_for_user(
        db,
        org_id=org_id,
        project_id=project_id,
        itinerary_id=itinerary_id,
        user=user,
        require_edit=False,
    )
    if not itinerary:
        return None

    return list(
        db.execute(
            select(ItineraryVersion)
            .where(ItineraryVersion.itinerary_id == itinerary_id)
            .options(selectinload(ItineraryVersion.created_by_user))
            .order_by(ItineraryVersion.version_number.desc())
        )
        .scalars()
        .all()
    )


def list_project_planner_users(db: Session, *, org_id: str, project_id: str) -> list[User]:
    return list(
        db.execute(
            select(User)
            .join(ProjectMember, ProjectMember.user_id == User.id)
            .where(User.org_id == org_id, ProjectMember.project_id == project_id, User.is_active.is_(True))
            .options(selectinload(User.user_roles).selectinload(UserRole.role))
            .order_by(User.username.asc())
        )
        .scalars()
        .all()
    )


def itinerary_day_count(db: Session, *, itinerary_id: str) -> int:
    return (
        db.execute(select(func.count(ItineraryDay.id)).where(ItineraryDay.itinerary_id == itinerary_id)).scalar_one()
    )


def _sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _json_bytes(payload: object) -> bytes:
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"), sort_keys=True).encode("utf-8")


def _sync_file_entry(path: str, kind: str, content: bytes) -> dict[str, object]:
    return {
        "path": path,
        "kind": kind,
        "sha256": _sha256_bytes(content),
        "bytes": len(content),
    }


def _serialize_itinerary_for_sync(itinerary: Itinerary) -> dict[str, object]:
    days_payload: list[dict[str, object]] = []
    for day in sorted(itinerary.days, key=lambda row: row.day_number):
        stops_payload: list[dict[str, object]] = []
        for stop in sorted(day.stops, key=lambda row: row.order_index):
            attraction = stop.attraction
            stops_payload.append(
                {
                    "id": stop.id,
                    "order_index": stop.order_index,
                    "start_minute_of_day": stop.start_minute_of_day,
                    "duration_minutes": stop.duration_minutes,
                    "notes": stop.notes,
                    "attraction_id": attraction.id,
                    "attraction_key": attraction.normalized_dedupe_key,
                    "attraction_name": attraction.name,
                    "attraction_city": attraction.city,
                    "attraction_state": attraction.state,
                }
            )

        days_payload.append(
            {
                "id": day.id,
                "day_number": day.day_number,
                "title": day.title,
                "notes": day.notes,
                "urban_speed_mph_override": day.urban_speed_mph_override,
                "highway_speed_mph_override": day.highway_speed_mph_override,
                "stops": stops_payload,
            }
        )

    source_version = max(itinerary.version_counter, 1)
    return {
        "record_type": "itinerary",
        "entity_id": itinerary.id,
        "entity_name": itinerary.name,
        "base_version": source_version,
        "target_version": source_version + 1,
        "payload": {
            "name": itinerary.name,
            "description": itinerary.description,
            "status": itinerary.status,
            "assigned_planner_username": itinerary.assigned_planner.username if itinerary.assigned_planner else None,
            "urban_speed_mph_override": itinerary.urban_speed_mph_override,
            "highway_speed_mph_override": itinerary.highway_speed_mph_override,
            "days": days_payload,
        },
    }


def export_sync_package_archive(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    user: User,
) -> tuple[bytes, str] | None:
    project = _project_for_user(db, org_id=org_id, project_id=project_id, user=user, require_edit=False)
    if not project:
        return None

    org = db.execute(select(Organization).where(Organization.id == org_id)).scalars().one()
    itineraries = list(
        db.execute(
            select(Itinerary)
            .where(Itinerary.org_id == org_id, Itinerary.project_id == project_id)
            .options(
                selectinload(Itinerary.assigned_planner),
                selectinload(Itinerary.days)
                .selectinload(ItineraryDay.stops)
                .selectinload(ItineraryStop.attraction)
                .selectinload(Attraction.dataset),
            )
            .order_by(Itinerary.name.asc())
        )
        .scalars()
        .all()
    )

    serialized_records = [_serialize_itinerary_for_sync(itinerary) for itinerary in itineraries]
    referenced_attractions: dict[str, dict[str, object]] = {}
    for itinerary in itineraries:
        for day in itinerary.days:
            for stop in day.stops:
                attraction = stop.attraction
                if attraction.id in referenced_attractions:
                    continue
                referenced_attractions[attraction.id] = {
                    "id": attraction.id,
                    "normalized_dedupe_key": attraction.normalized_dedupe_key,
                    "name": attraction.name,
                    "city": attraction.city,
                    "state": attraction.state,
                    "dataset_name": attraction.dataset.name if attraction.dataset else None,
                }

    data_payload = {
        "project": {
            "project_id": project.id,
            "project_code": project.code,
            "project_name": project.name,
        },
        "records": serialized_records,
    }
    assets_payload = {"attractions": list(referenced_attractions.values())}

    data_bytes = _json_bytes(data_payload)
    assets_bytes = _json_bytes(assets_payload)
    manifest = {
        "package_type": SYNC_PACKAGE_TYPE,
        "format_version": SYNC_PACKAGE_FORMAT_VERSION,
        "package_id": str(uuid4()),
        "generated_at": datetime.now(UTC).isoformat(),
        "source": {
            "org_id": org_id,
            "org_slug": org.slug,
            "project_id": project.id,
            "project_code": project.code,
            "project_name": project.name,
            "exported_by_user_id": user.id,
            "exported_by_username": user.username,
        },
        "counts": {
            "records": len(serialized_records),
            "referenced_attractions": len(referenced_attractions),
        },
        "files": [
            _sync_file_entry(SYNC_DATA_PATH, "data", data_bytes),
            _sync_file_entry(SYNC_ASSETS_PATH, "asset", assets_bytes),
        ],
    }
    manifest_bytes = _json_bytes(manifest)

    package_buffer = io.BytesIO()
    with zipfile.ZipFile(package_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(SYNC_MANIFEST_PATH, manifest_bytes)
        archive.writestr(SYNC_DATA_PATH, data_bytes)
        archive.writestr(SYNC_ASSETS_PATH, assets_bytes)

    return package_buffer.getvalue(), f"{project.code}_sync_package"


def _load_sync_archive_entries(content: bytes) -> tuple[dict[str, bytes], list[str]]:
    file_errors: list[str] = []
    settings = get_settings()
    try:
        infos, _ = _inspect_zip_payload(
            content,
            max_entries=settings.planner_sync_package_max_entries,
            max_uncompressed_bytes=settings.planner_sync_package_max_uncompressed_bytes,
        )
        with zipfile.ZipFile(io.BytesIO(content), mode="r") as archive:
            entries: dict[str, bytes] = {}
            for info in infos:
                path = info.filename
                if path.startswith("/") or ".." in path.split("/"):
                    file_errors.append(f"Archive entry path is unsafe: {path}")
                    continue
                entries[path] = archive.read(path)
            return entries, file_errors
    except PlannerValidationError as exc:
        raise PlannerValidationError(str(exc)) from exc


def _coerce_optional_float(value: object, *, field_name: str) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        parsed = float(value)
    elif isinstance(value, str):
        parsed = float(value)
    else:
        raise PlannerValidationError(f"{field_name} must be a number")

    if parsed <= 0:
        raise PlannerValidationError(f"{field_name} must be positive")
    return parsed


def _resolve_sync_attraction_id(
    *,
    stop_payload: dict[str, object],
    catalog_by_id: dict[str, Attraction],
    catalog_by_key: dict[str, Attraction],
    attraction_key_by_id: dict[str, str],
) -> str:
    incoming_id = str(stop_payload.get("attraction_id") or "").strip()
    if incoming_id and incoming_id in catalog_by_id:
        return incoming_id

    dedupe_key = str(stop_payload.get("attraction_key") or "").strip()
    if not dedupe_key and incoming_id:
        dedupe_key = attraction_key_by_id.get(incoming_id, "")
    if dedupe_key and dedupe_key in catalog_by_key:
        return catalog_by_key[dedupe_key].id

    stop_name = str(stop_payload.get("attraction_name") or "unknown")
    raise PlannerValidationError(
        f"Stop attraction '{stop_name}' is not available in this project's linked active catalog"
    )


def _validate_sync_days_payload(days_payload: object) -> list[dict[str, object]]:
    if not isinstance(days_payload, list):
        raise PlannerValidationError("Record payload days must be a list")
    normalized: list[dict[str, object]] = []
    seen_day_numbers: set[int] = set()
    for day_row in days_payload:
        if not isinstance(day_row, dict):
            raise PlannerValidationError("Each day payload must be an object")
        day_number_raw = day_row.get("day_number")
        if not isinstance(day_number_raw, int) or day_number_raw < 1:
            raise PlannerValidationError("Day number must be a positive integer")
        if day_number_raw in seen_day_numbers:
            raise PlannerValidationError("Day number must be unique within a synced itinerary")
        seen_day_numbers.add(day_number_raw)
        normalized.append(day_row)
    return sorted(normalized, key=lambda row: int(row["day_number"]))


def _apply_sync_days_and_stops(
    db: Session,
    *,
    itinerary: Itinerary,
    days_payload: object,
    catalog_by_id: dict[str, Attraction],
    catalog_by_key: dict[str, Attraction],
    attraction_key_by_id: dict[str, str],
) -> tuple[int, int]:
    for day in list(itinerary.days):
        db.delete(day)
    db.flush()

    day_count = 0
    stop_count = 0
    for day_row in _validate_sync_days_payload(days_payload):
        day = ItineraryDay(
            id=str(day_row.get("id") or str(uuid4())),
            itinerary_id=itinerary.id,
            day_number=int(day_row["day_number"]),
            title=(str(day_row.get("title")) if day_row.get("title") is not None else None),
            notes=(str(day_row.get("notes")) if day_row.get("notes") is not None else None),
            urban_speed_mph_override=_coerce_optional_float(
                day_row.get("urban_speed_mph_override"), field_name="Day urban speed override"
            ),
            highway_speed_mph_override=_coerce_optional_float(
                day_row.get("highway_speed_mph_override"), field_name="Day highway speed override"
            ),
        )
        db.add(day)
        db.flush()
        day_count += 1

        stops_payload = day_row.get("stops", [])
        if not isinstance(stops_payload, list):
            raise PlannerValidationError("Day stops payload must be a list")

        ordered_stops = []
        for row in stops_payload:
            if not isinstance(row, dict):
                raise PlannerValidationError("Stop payload rows must be objects")
            if not isinstance(row.get("start_minute_of_day"), int):
                raise PlannerValidationError("Stop start minute must be an integer")
            if not isinstance(row.get("duration_minutes"), int):
                raise PlannerValidationError("Stop duration must be an integer")
            ordered_stops.append(row)

        ordered_stops.sort(key=lambda row: int(row.get("order_index", 0)))
        for order_index, stop_row in enumerate(ordered_stops):
            attraction_id = _resolve_sync_attraction_id(
                stop_payload=stop_row,
                catalog_by_id=catalog_by_id,
                catalog_by_key=catalog_by_key,
                attraction_key_by_id=attraction_key_by_id,
            )
            stop = ItineraryStop(
                id=str(stop_row.get("id") or str(uuid4())),
                itinerary_day_id=day.id,
                attraction_id=attraction_id,
                order_index=order_index,
                start_minute_of_day=int(stop_row["start_minute_of_day"]),
                duration_minutes=int(stop_row["duration_minutes"]),
                notes=(str(stop_row.get("notes")) if stop_row.get("notes") is not None else None),
            )
            db.add(stop)
            stop_count += 1

    db.flush()
    return day_count, stop_count


def _sync_planner_assignment(
    *,
    planner_user_by_username: dict[str, str],
    assigned_planner_username: object,
) -> tuple[str | None, str | None]:
    if not isinstance(assigned_planner_username, str) or not assigned_planner_username.strip():
        return None, None
    planner_id = planner_user_by_username.get(assigned_planner_username)
    if planner_id:
        return planner_id, None
    return None, f"Assigned planner '{assigned_planner_username}' was not found in destination project membership"


def _create_sync_version_row(
    db: Session,
    *,
    itinerary: Itinerary,
    org: Organization,
    changed_by_user_id: str,
    version_number: int,
    change_summary: str,
) -> None:
    snapshot_itinerary = _itinerary_with_graph(db, itinerary_id=itinerary.id)
    if not snapshot_itinerary:
        raise PlannerValidationError("Failed to resolve itinerary snapshot during sync import")
    db.add(
        ItineraryVersion(
            org_id=itinerary.org_id,
            project_id=itinerary.project_id,
            itinerary_id=itinerary.id,
            version_number=version_number,
            change_summary=change_summary,
            snapshot=_serialize_snapshot(snapshot_itinerary, org),
            created_by_user_id=changed_by_user_id,
        )
    )
    db.flush()


def import_sync_package_archive(
    db: Session,
    *,
    org_id: str,
    project_id: str,
    user: User,
    file_name: str,
    content: bytes,
) -> dict | None:
    project = _project_for_user(db, org_id=org_id, project_id=project_id, user=user, require_edit=True)
    if not project:
        return None

    if not file_name.lower().endswith(".zip"):
        raise PlannerValidationError("Sync package upload must be a .zip file.")

    org = db.execute(select(Organization).where(Organization.id == org_id)).scalars().one()
    receipt: dict[str, object] = {
        "project_id": project.id,
        "file_name": file_name,
        "imported_at": datetime.now(UTC),
        "format_version": None,
        "integrity_validated": False,
        "total_record_count": 0,
        "inserted_record_count": 0,
        "updated_record_count": 0,
        "conflict_count": 0,
        "rejected_record_count": 0,
        "applied_record_count": 0,
        "file_errors": [],
        "correction_hints": [],
        "record_results": [],
    }

    def finalize_receipt() -> dict[str, object]:
        file_errors = [str(message) for message in receipt.get("file_errors", []) if isinstance(message, str)]
        record_results = [row for row in receipt.get("record_results", []) if isinstance(row, dict)]
        hints: list[str] = []

        if any("Checksum mismatch" in message for message in file_errors):
            hints.append("Re-export or retransmit the sync package to restore file integrity before importing again.")
        if any(
            marker in message
            for message in file_errors
            for marker in (
                "Manifest",
                "Data file",
                "Assets file",
                "Archive file missing",
                "Archive entry path is unsafe",
            )
        ):
            hints.append("Create a fresh sync package export from TrailForge before retrying this import.")
        if any(
            marker in message
            for message in file_errors
            for marker in (
                "organization does not match",
                "project code does not match",
            )
        ):
            hints.append("Import the sync package into the matching organization and selected project.")
        if any(row.get("action") == "conflict" for row in record_results):
            hints.append("Refresh the destination itinerary state, reconcile local edits, and re-export before importing again.")
        if any(row.get("action") == "rejected" for row in record_results):
            hints.append("Review rejected records in the receipt and correct the source itinerary data before the next import.")

        receipt["correction_hints"] = hints
        return receipt

    entries, archive_errors = _load_sync_archive_entries(content)
    if archive_errors:
        receipt["file_errors"] = archive_errors
        return finalize_receipt()

    file_errors: list[str] = []
    manifest_bytes = entries.get(SYNC_MANIFEST_PATH)
    if not manifest_bytes:
        file_errors.append("Manifest file is missing from archive")
        receipt["file_errors"] = file_errors
        return finalize_receipt()

    try:
        manifest = json.loads(manifest_bytes.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError):
        receipt["file_errors"] = ["Manifest file is not valid UTF-8 JSON"]
        return finalize_receipt()

    if not isinstance(manifest, dict):
        receipt["file_errors"] = ["Manifest root must be a JSON object"]
        return finalize_receipt()

    receipt["format_version"] = manifest.get("format_version") if isinstance(manifest.get("format_version"), str) else None

    if manifest.get("package_type") != SYNC_PACKAGE_TYPE:
        file_errors.append("Manifest package_type does not match TrailForge sync package type")
    if manifest.get("format_version") != SYNC_PACKAGE_FORMAT_VERSION:
        file_errors.append("Manifest format_version is not supported")

    source = manifest.get("source")
    if not isinstance(source, dict):
        file_errors.append("Manifest source section is missing")
    else:
        if source.get("org_slug") != org.slug:
            file_errors.append("Package source organization does not match current organization")
        if source.get("project_code") != project.code:
            file_errors.append("Package source project code does not match selected project")

    manifest_files = manifest.get("files")
    if not isinstance(manifest_files, list) or not manifest_files:
        file_errors.append("Manifest files list is missing or empty")
    else:
        for entry in manifest_files:
            if not isinstance(entry, dict):
                file_errors.append("Manifest files entries must be objects")
                continue
            path = entry.get("path")
            expected_sha = entry.get("sha256")
            if not isinstance(path, str) or not path:
                file_errors.append("Manifest file entry missing path")
                continue
            if not isinstance(expected_sha, str) or not expected_sha:
                file_errors.append(f"Manifest checksum missing for {path}")
                continue
            content_bytes = entries.get(path)
            if content_bytes is None:
                file_errors.append(f"Archive file missing: {path}")
                continue
            if _sha256_bytes(content_bytes) != expected_sha:
                file_errors.append(f"Checksum mismatch for {path}")

    if SYNC_DATA_PATH not in entries:
        file_errors.append("Data file missing from archive")
    if SYNC_ASSETS_PATH not in entries:
        file_errors.append("Assets file missing from archive")

    if file_errors:
        receipt["file_errors"] = file_errors
        return finalize_receipt()

    receipt["integrity_validated"] = True

    try:
        data_payload = json.loads(entries[SYNC_DATA_PATH].decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError):
        receipt["file_errors"] = ["Data file is not valid UTF-8 JSON"]
        receipt["integrity_validated"] = False
        return finalize_receipt()

    try:
        assets_payload = json.loads(entries[SYNC_ASSETS_PATH].decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError):
        receipt["file_errors"] = ["Assets file is not valid UTF-8 JSON"]
        receipt["integrity_validated"] = False
        return finalize_receipt()

    if not isinstance(data_payload, dict):
        receipt["file_errors"] = ["Data payload root must be a JSON object"]
        receipt["integrity_validated"] = False
        return finalize_receipt()

    records = data_payload.get("records")
    if not isinstance(records, list):
        receipt["file_errors"] = ["Data payload records must be an array"]
        receipt["integrity_validated"] = False
        return finalize_receipt()

    attraction_key_by_id: dict[str, str] = {}
    if isinstance(assets_payload, dict) and isinstance(assets_payload.get("attractions"), list):
        for row in assets_payload["attractions"]:
            if not isinstance(row, dict):
                continue
            attraction_id = row.get("id")
            dedupe_key = row.get("normalized_dedupe_key")
            if isinstance(attraction_id, str) and isinstance(dedupe_key, str):
                attraction_key_by_id[attraction_id] = dedupe_key

    catalog_by_id = _project_catalog_map(db, org_id=org_id, project_id=project_id)
    catalog_by_key: dict[str, Attraction] = {}
    ambiguous_keys: set[str] = set()
    for attraction in catalog_by_id.values():
        key = attraction.normalized_dedupe_key
        if key in catalog_by_key:
            ambiguous_keys.add(key)
            continue
        catalog_by_key[key] = attraction
    for key in ambiguous_keys:
        catalog_by_key.pop(key, None)

    planner_user_by_username = {
        row.username: row.id
        for row in list_project_planner_users(db, org_id=org_id, project_id=project_id)
        if _is_planner(db, row)
    }

    record_results: list[dict[str, object]] = []
    inserted = 0
    updated = 0
    conflicts = 0
    rejected = 0

    for raw_record in records:
        if not isinstance(raw_record, dict):
            rejected += 1
            record_results.append(
                {
                    "record_type": "unknown",
                    "entity_id": "unknown",
                    "entity_name": "unknown",
                    "action": "rejected",
                    "base_version": 0,
                    "target_version": 0,
                    "destination_version": None,
                    "message": "Record entry must be an object",
                }
            )
            continue

        record_type = str(raw_record.get("record_type") or "")
        entity_id = str(raw_record.get("entity_id") or "")
        entity_name = str(raw_record.get("entity_name") or "")
        base_version = raw_record.get("base_version")
        target_version = raw_record.get("target_version")
        payload = raw_record.get("payload")

        if record_type != "itinerary":
            rejected += 1
            record_results.append(
                {
                    "record_type": record_type or "unknown",
                    "entity_id": entity_id or "unknown",
                    "entity_name": entity_name or "unknown",
                    "action": "rejected",
                    "base_version": int(base_version) if isinstance(base_version, int) else 0,
                    "target_version": int(target_version) if isinstance(target_version, int) else 0,
                    "destination_version": None,
                    "message": "Unsupported record_type in sync package",
                }
            )
            continue

        if not isinstance(base_version, int) or not isinstance(target_version, int):
            rejected += 1
            record_results.append(
                {
                    "record_type": record_type,
                    "entity_id": entity_id or "unknown",
                    "entity_name": entity_name or "unknown",
                    "action": "rejected",
                    "base_version": 0,
                    "target_version": 0,
                    "destination_version": None,
                    "message": "Sync record versions must be integers",
                }
            )
            continue

        if target_version <= base_version:
            rejected += 1
            record_results.append(
                {
                    "record_type": record_type,
                    "entity_id": entity_id or "unknown",
                    "entity_name": entity_name or "unknown",
                    "action": "rejected",
                    "base_version": base_version,
                    "target_version": target_version,
                    "destination_version": None,
                    "message": "target_version must be greater than base_version",
                }
            )
            continue

        if not isinstance(payload, dict):
            rejected += 1
            record_results.append(
                {
                    "record_type": record_type,
                    "entity_id": entity_id or "unknown",
                    "entity_name": entity_name or "unknown",
                    "action": "rejected",
                    "base_version": base_version,
                    "target_version": target_version,
                    "destination_version": None,
                    "message": "Record payload must be an object",
                }
            )
            continue

        existing = (
            db.execute(
                select(Itinerary)
                .where(
                    Itinerary.id == entity_id,
                    Itinerary.org_id == org_id,
                    Itinerary.project_id == project_id,
                )
                .options(selectinload(Itinerary.days).selectinload(ItineraryDay.stops))
            )
            .scalars()
            .first()
        )

        if existing and existing.version_counter != base_version:
            conflicts += 1
            record_results.append(
                {
                    "record_type": record_type,
                    "entity_id": entity_id,
                    "entity_name": existing.name,
                    "action": "conflict",
                    "base_version": base_version,
                    "target_version": target_version,
                    "destination_version": existing.version_counter,
                    "message": "Destination version does not match incoming base version; manual reconciliation required",
                }
            )
            continue

        try:
            assigned_planner_id, assignment_message = _sync_planner_assignment(
                planner_user_by_username=planner_user_by_username,
                assigned_planner_username=payload.get("assigned_planner_username"),
            )
            itinerary_name = str(payload.get("name") or entity_name or "")
            if not itinerary_name:
                raise PlannerValidationError("Itinerary name is required in sync payload")

            if existing is None:
                itinerary = Itinerary(
                    id=entity_id or str(uuid4()),
                    org_id=org_id,
                    project_id=project_id,
                    name=itinerary_name,
                    description=str(payload.get("description")) if payload.get("description") is not None else None,
                    status=str(payload.get("status") or "draft"),
                    assigned_planner_user_id=assigned_planner_id,
                    urban_speed_mph_override=_coerce_optional_float(
                        payload.get("urban_speed_mph_override"), field_name="Itinerary urban speed override"
                    ),
                    highway_speed_mph_override=_coerce_optional_float(
                        payload.get("highway_speed_mph_override"), field_name="Itinerary highway speed override"
                    ),
                    version_counter=target_version,
                    created_by_user_id=user.id,
                    updated_by_user_id=user.id,
                )
                db.add(itinerary)
                db.flush()
                day_count, stop_count = _apply_sync_days_and_stops(
                    db,
                    itinerary=itinerary,
                    days_payload=payload.get("days", []),
                    catalog_by_id=catalog_by_id,
                    catalog_by_key=catalog_by_key,
                    attraction_key_by_id=attraction_key_by_id,
                )
                _create_sync_version_row(
                    db,
                    itinerary=itinerary,
                    org=org,
                    changed_by_user_id=user.id,
                    version_number=target_version,
                    change_summary="Sync package record inserted",
                )
                db.commit()
                inserted += 1
                record_results.append(
                    {
                        "record_type": record_type,
                        "entity_id": itinerary.id,
                        "entity_name": itinerary.name,
                        "action": "inserted",
                        "base_version": base_version,
                        "target_version": target_version,
                        "destination_version": target_version,
                        "message": assignment_message
                        or f"Inserted itinerary with {day_count} day(s) and {stop_count} stop(s)",
                    }
                )
            else:
                existing.name = itinerary_name
                existing.description = str(payload.get("description")) if payload.get("description") is not None else None
                existing.status = str(payload.get("status") or "draft")
                existing.assigned_planner_user_id = assigned_planner_id
                existing.urban_speed_mph_override = _coerce_optional_float(
                    payload.get("urban_speed_mph_override"), field_name="Itinerary urban speed override"
                )
                existing.highway_speed_mph_override = _coerce_optional_float(
                    payload.get("highway_speed_mph_override"), field_name="Itinerary highway speed override"
                )
                existing.version_counter = target_version
                existing.updated_by_user_id = user.id
                db.add(existing)
                day_count, stop_count = _apply_sync_days_and_stops(
                    db,
                    itinerary=existing,
                    days_payload=payload.get("days", []),
                    catalog_by_id=catalog_by_id,
                    catalog_by_key=catalog_by_key,
                    attraction_key_by_id=attraction_key_by_id,
                )
                _create_sync_version_row(
                    db,
                    itinerary=existing,
                    org=org,
                    changed_by_user_id=user.id,
                    version_number=target_version,
                    change_summary="Sync package record updated",
                )
                db.commit()
                updated += 1
                record_results.append(
                    {
                        "record_type": record_type,
                        "entity_id": existing.id,
                        "entity_name": existing.name,
                        "action": "updated",
                        "base_version": base_version,
                        "target_version": target_version,
                        "destination_version": target_version,
                        "message": assignment_message
                        or f"Updated itinerary with {day_count} day(s) and {stop_count} stop(s)",
                    }
                )
        except (PlannerValidationError, IntegrityError) as exc:
            db.rollback()
            rejected += 1
            record_results.append(
                {
                    "record_type": record_type,
                    "entity_id": entity_id or "unknown",
                    "entity_name": entity_name or "unknown",
                    "action": "rejected",
                    "base_version": base_version,
                    "target_version": target_version,
                    "destination_version": existing.version_counter if existing else None,
                    "message": str(exc),
                }
            )

    total_records = len(records)
    receipt["total_record_count"] = total_records
    receipt["inserted_record_count"] = inserted
    receipt["updated_record_count"] = updated
    receipt["conflict_count"] = conflicts
    receipt["rejected_record_count"] = rejected
    receipt["applied_record_count"] = inserted + updated
    receipt["record_results"] = record_results
    return finalize_receipt()
